from .models import Proposal
from .models import Broker
from .models import SamsungCode
from .models import Product
from .models import Customer
from openpyxl import load_workbook
from openpyxl import Workbook

def save(request):

    print("proposal_save")

    for i in range(0,1):

        contact_conclusion_date = None; samsung_code = None; samsung_sales_manager = None;  
        team = None;  sales_manager = None;  company_registration_number = None
        address = None; email = None; resident_registration_number = None; team = None; rate = None; decision_price = None; sales_price = None

        contact_conclusion_date = request.POST.get('contact_conclusion_date' + str(i), None)
        if contact_conclusion_date == '' : 
            contact_conclusion_date = None

        samsung_code = request.POST.get('samsung_code'+ str(i), None)
        if samsung_code == '' : 
            samsung_code = None
        else :
            count_samsung_code_data = SamsungCode.objects.values('samsung_code').filter(samsung_code=samsung_code).count()
            print("count_samsung_code_data")
            print(count_samsung_code_data)
            if count_samsung_code_data == 0:
                return "등록되지 않은 삼성코드입니다."

        samsung_sales_manager = request.POST.get('samsung_sales_manager'+ str(i), None)
        print("samsung_sales_manager")
        print(samsung_sales_manager)
        if samsung_sales_manager == '' : 
            samsung_sales_manager = None
        else :
            count_samsung_sales_manager_data = SamsungCode.objects.values('manager').filter(manager=samsung_sales_manager).count()
            print("count_samsung_sales_manager_data")
            print(count_samsung_sales_manager_data)
            if count_samsung_sales_manager_data == 0:
                return "등록되지 않은 삼성영업담당입니다."

        team = request.POST.get('team' + str(i), None)
        if team == '' : 
            team = None

        sales_manager = request.POST.get('sales_manager'+ str(i), None)
        count_sales_manager_data = SamsungCode.objects.values('manager').filter(manager=sales_manager).count()
        print(count_sales_manager_data)
        if sales_manager == '' : 
            sales_manager = None
        elif count_sales_manager_data == 0:
            return "등록되지 않은 영업담당입니다."

        broker = request.POST.get('broker'+ str(i), None)
        count_broker_data = Broker.objects.values('name').filter(name=broker).count()
        print(count_broker_data)
        if broker == '' : 
            broker = None
        elif count_broker_data == 0:
            return "등록되지 않은 중개사입니다. 중개사를 등록해주세요."

        scm_manager = request.POST.get('scm_manager'+ str(i), None)
        count_scm_manager_data = SamsungCode.objects.values('manager').filter(manager=scm_manager).count()
        print(count_scm_manager_data)
        if scm_manager == '' : 
            scm_manager = None
        elif count_scm_manager_data == 0:
            return "등록되지 않은 SCM담당입니다."

        customer_name = request.POST.get('customer_name'+ str(i), None)
        count_customer_name_data = Customer.objects.values('customer_name').filter(customer_name=customer_name).count()
        print(count_customer_name_data)
        if customer_name == '' : 
            customer_name = None
        elif count_customer_name_data == 0:
            return "등록되지 않은 업체입니다. 업체를 등록해주세요."

        company_registration_number = request.POST.get('company_registration_number'+ str(i), None)
        # customer_name인 company_registration_number가 있어야 함
        count_company_registration_number_data = Customer.objects.values('customer_name').filter(company_registration_number=company_registration_number).filter(customer_name=customer_name).count()
        print(count_company_registration_number_data)
        if company_registration_number == '' : 
            company_registration_number = None
        elif count_company_registration_number_data == 0:
            return "등록된 사업자 등록번호와 업체명이 일치하지 않습니다"

        payment_method = request.POST.get('payment_method'+ str(i), None)
        if payment_method == '' : 
            payment_method = None
        sales_type = request.POST.get('sales_type'+ str(i), None)
        if sales_type == '' : 
            sales_type = None
        demand = request.POST.get('demand'+ str(i), None)
        if demand == '' : 
            demand = None
        billing_place = request.POST.get('billing_place'+ str(i), None)
        if billing_place == '' : 
            billing_place = None
        oppty_num = request.POST.get('oppty_num'+ str(i), None)
        if oppty_num == '' : 
            oppty_num = None
        productno = request.POST.get('productno'+ str(i), None)
        count_productno_data = Product.objects.values('productno').filter(productno=productno).count()
        print(count_productno_data)
        if productno == '' : 
            productno = None
        elif count_productno_data == 0:
            return "등록되지 않은 productno입니다."

        buy_place = request.POST.get('buy_place'+ str(i), None)
        if buy_place == '' : 
            buy_place = None
        decision_quantity = request.POST.get('decision_quantity'+ str(i), None)
        decision_unit = request.POST.get('decision_unit'+ str(i), None)
        sales_unit = request.POST.get('sales_unit'+ str(i), None)
        
        # 자동 계산 decision_quantity * decision_unit
        # 자동 계산 decision_quantity * sales_unit
        if decision_quantity == '' and sales_unit != '' :
            sales_price = sales_unit
        elif decision_quantity == '' and decision_unit != '' : 
            decision_price = decision_unit
        elif decision_quantity != '' and sales_unit == '' : 
            sales_price = decision_quantity
        elif decision_quantity != '' and decision_unit == '' : 
            decision_price = decision_quantity
        elif decision_quantity == '' and sales_unit == '' : 
            sales_price = None
        elif decision_quantity == '' and decision_unit == '' : 
            decision_price = None
        else:
            decision_price = int(decision_quantity) * int(decision_unit)
            sales_price = int(decision_quantity) * int(sales_unit)
        
        if decision_quantity == '' : 
            decision_quantity = None
        if decision_unit == '' : 
            decision_unit = None
        if sales_unit == '' : 
            sales_unit = None

        delivery_request_date = request.POST.get('delivery_request_date'+ str(i), None)
        if delivery_request_date == '' : 
            delivery_request_date = None
        recipient = request.POST.get('recipient'+ str(i), None)
        if recipient == '' : 
            recipient = None
        recipient_phone1 = request.POST.get('recipient_phone1'+ str(i), None)
        if recipient_phone1 == '' : 
            recipient_phone1 = None
        recipient_phone2 = request.POST.get('recipient_phone2'+ str(i), None)
        if recipient_phone2 == '' : 
            recipient_phone2 = None
        delivery_address = request.POST.get('delivery_address'+ str(i), None)
        if delivery_address == '' : 
            delivery_address = None

        if oppty_num != '' and oppty_num != None:
            proposal_data = Proposal( contact_conclusion_date, samsung_code, samsung_sales_manager, team, sales_manager, broker, scm_manager,  
                                    customer_name, company_registration_number, payment_method, sales_type, demand, billing_place, oppty_num, productno, buy_place,  decision_quantity,
                                    decision_unit, decision_price, sales_unit, sales_price, delivery_request_date, recipient, 
                                    recipient_phone1, recipient_phone2, delivery_address) 
            proposal_data.save()
            isSuccess = "저장되었습니다"
        else :
            print("oppty_num")
            isSuccess = "저장에 실패했습니다"
    return isSuccess

def select(request):
    select_oppty_num = request.POST.get('select_oppty_num', None)
    print("select = ")
    print(select_oppty_num)
    if select_oppty_num == '' : 
            select_oppty_num = None
    return select_oppty_num

# def upload(request):
#     isSuccess = "실패"
#     if request.method == 'POST':
#         if 'file' in request.FILES:
#             wb = load_workbook(filename=request.FILES['file'].file)
#             first_sheet = wb.get_sheet_names()[0]
#             worksheet = wb.get_sheet_by_name(first_sheet)
#             print(worksheet)
            
#             for row in worksheet.iter_rows(min_row=2): # Offset for header
#                 employee = Employee()
#                 employee.name = row[0].value
#                 employee.team = row[1].value
#                 employee.position = row[2].value
#                 employee.department = row[3].value
#                 employee.resident_registration_number = row[4].value
#                 employee.rate = row[5].value
#                 employee.phone_num = row[6].value
#                 employee.office_num = row[7].value
#                 employee.recipient = row[8].value
#                 employee.address = row[9].value
#                 employee.email = row[10].value
#                 employee.charge = row[11].value
#                 employee.passwd = row[12].value
#                 employee.authority = row[13].value

#                 if employee.name != '' and employee.name != None:
#                     employee.save()
#                     isSuccess = "성공"
#     return isSuccess

def modify(request):

    print("proposal_save")

    for i in range(0,1):

        contact_conclusion_date = None; samsung_code = None; samsung_sales_manager = None;  
        team = None;  sales_manager = None;  
        address = None; email = None; resident_registration_number = None; team = None; rate = None; decision_price = None; sales_price = None

        contact_conclusion_date = request.POST.get('modify_contact_conclusion_date' + str(i), None)
        if contact_conclusion_date == '' : 
            contact_conclusion_date = None

        samsung_code = request.POST.get('modify_samsung_code'+ str(i), None)
        if samsung_code == '' : 
            samsung_code = None
        else :
            count_samsung_code_data = SamsungCode.objects.values('samsung_code').filter(samsung_code=samsung_code).count()
            print("count_samsung_code_data")
            print(count_samsung_code_data)
            if count_samsung_code_data == 0:
                return "등록되지 않은 삼성코드입니다."

        samsung_sales_manager = request.POST.get('modify_samsung_sales_manager'+ str(i), None)
        print("samsung_sales_manager")
        print(samsung_sales_manager)
        if samsung_sales_manager == '' : 
            samsung_sales_manager = None
        else :
            count_samsung_sales_manager_data = SamsungCode.objects.values('manager').filter(manager=samsung_sales_manager).count()
            print("count_samsung_sales_manager_data")
            print(count_samsung_sales_manager_data)
            if count_samsung_sales_manager_data == 0:
                return "등록되지 않은 삼성영업담당입니다."

        team = request.POST.get('modify_team' + str(i), None)
        if team == '' : 
            team = None

        sales_manager = request.POST.get('modify_sales_manager'+ str(i), None)
        count_sales_manager_data = SamsungCode.objects.values('manager').filter(manager=sales_manager).count()
        print(count_sales_manager_data)
        if sales_manager == '' : 
            sales_manager = None
        elif count_sales_manager_data == 0:
            return "등록되지 않은 영업담당입니다."

        broker = request.POST.get('modify_broker'+ str(i), None)
        count_broker_data = Broker.objects.values('name').filter(name=broker).count()
        print(count_broker_data)
        if broker == '' : 
            broker = None
        elif count_broker_data == 0:
            return "등록되지 않은 중개사입니다. 중개사를 등록해주세요."

        scm_manager = request.POST.get('modify_scm_manager'+ str(i), None)
        count_scm_manager_data = SamsungCode.objects.values('manager').filter(manager=scm_manager).count()
        print(count_scm_manager_data)
        if scm_manager == '' : 
            scm_manager = None
        elif count_scm_manager_data == 0:
            return "등록되지 않은 SCM담당입니다."

        customer_name = request.POST.get('modify_customer_name'+ str(i), None)
        count_customer_name_data = Customer.objects.values('customer_name').filter(customer_name=customer_name).count()
        print(count_customer_name_data)
        if customer_name == '' : 
            customer_name = None
        elif count_customer_name_data == 0:
            return "등록되지 않은 업체입니다. 업체를 등록해주세요."

        company_registration_number = request.POST.get('modify_company_registration_number'+ str(i), None)
        # customer_name인 company_registration_number가 있어야 함
        count_company_registration_number_data = Customer.objects.values('customer_name').filter(company_registration_number=company_registration_number).filter(customer_name=customer_name).count()
        print(count_company_registration_number_data)
        if company_registration_number == '' : 
            company_registration_number = None
        elif count_company_registration_number_data == 0:
            return "등록된 사업자 등록번호와 업체명이 일치하지 않습니다"

        payment_method = request.POST.get('modify_payment_method'+ str(i), None)
        if payment_method == '' : 
            payment_method = None
        sales_type = request.POST.get('modify_sales_type'+ str(i), None)
        if sales_type == '' : 
            sales_type = None
        demand = request.POST.get('modify_demand'+ str(i), None)
        if demand == '' : 
            demand = None
        billing_place = request.POST.get('modify_billing_place'+ str(i), None)
        if billing_place == '' : 
            billing_place = None
        oppty_num = request.POST.get('modify_oppty_num'+ str(i), None)
        if oppty_num == '' : 
            oppty_num = None

        productno = request.POST.get('modify_productno'+ str(i), None)
        count_productno_data = Product.objects.values('productno').filter(productno=productno).count()
        print(count_productno_data)
        if productno == '' : 
            productno = None
        elif count_productno_data == 0:
            return "등록되지 않은 productno입니다."

        buy_place = request.POST.get('modify_buy_place'+ str(i), None)
        if buy_place == '' : 
            buy_place = None
        decision_quantity = request.POST.get('modify_decision_quantity'+ str(i), None)
        decision_unit = request.POST.get('modify_decision_unit'+ str(i), None)
        sales_unit = request.POST.get('modify_sales_unit'+ str(i), None)
        
        # 자동 계산 decision_quantity * decision_unit
        # 자동 계산 decision_quantity * sales_unit
        if decision_quantity == '' and sales_unit != '' :
            sales_price = sales_unit
        elif decision_quantity == '' and decision_unit != '' : 
            decision_price = decision_unit
        elif decision_quantity != '' and sales_unit == '' : 
            sales_price = decision_quantity
        elif decision_quantity != '' and decision_unit == '' : 
            decision_price = decision_quantity
        elif decision_quantity == '' and sales_unit == '' : 
            sales_price = None
        elif decision_quantity == '' and decision_unit == '' : 
            decision_price = None
        else:
            decision_price = int(decision_quantity) * int(decision_unit)
            sales_price = int(decision_quantity) * int(sales_unit)
        
        if decision_quantity == '' : 
            decision_quantity = None
        if decision_unit == '' : 
            decision_unit = None
        if sales_unit == '' : 
            sales_unit = None

        delivery_request_date = request.POST.get('modify_delivery_request_date'+ str(i), None)
        if delivery_request_date == '' : 
            delivery_request_date = None
        recipient = request.POST.get('modify_recipient'+ str(i), None)
        if recipient == '' : 
            recipient = None
        recipient_phone1 = request.POST.get('modify_recipient_phone1'+ str(i), None)
        if recipient_phone1 == '' : 
            recipient_phone1 = None
        recipient_phone2 = request.POST.get('modify_recipient_phone2'+ str(i), None)
        if recipient_phone2 == '' : 
            recipient_phone2 = None
        delivery_address = request.POST.get('modify_delivery_address'+ str(i), None)
        if delivery_address == '' : 
            delivery_address = None

        if oppty_num != '' and oppty_num != None:
            proposal_data = Proposal( contact_conclusion_date, samsung_code, samsung_sales_manager, team, sales_manager, broker, scm_manager,  
                                    customer_name, company_registration_number, payment_method, sales_type, demand, billing_place, oppty_num, productno, buy_place,  decision_quantity,
                                    decision_unit, decision_price, sales_unit, sales_price, delivery_request_date, recipient, 
                                    recipient_phone1, recipient_phone2, delivery_address) 
            proposal_data.save()
            isSuccess = "수정되었습니다"
        else :
            print("oppty_num")
            isSuccess = "수정에 실패했습니다"
    return isSuccess