from .models import Product
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl

def save(request):
    print("product_save")

    for i in range(0,1):
        productno = None; main_category = None; middle_category = None
        sub_category = None; professionalism_fee_rate =''; potential_fee_rate = None
        additional_fee1 = None; additional_fee2 = None; additional_fee3 = None

        productno = request.POST.get('productno' + str(i), None)
        main_category = request.POST.get('main_category' + str(i), None)
        if main_category == '' : 
            main_category = None
        middle_category = request.POST.get('middle_category' + str(i), None)
        if middle_category == '' : 
            middle_category = None
        sub_category = request.POST.get('sub_category' + str(i), None)
        if sub_category == '' : 
            sub_category = None
        professionalism_fee_rate = request.POST.get('professionalism_fee_rate' + str(i), None)
        if professionalism_fee_rate == '' : 
            professionalism_fee_rate = None
        potential_fee_rate = request.POST.get('potential_fee_rate' + str(i), None)
        if potential_fee_rate == '' : 
            potential_fee_rate = None
        additional_fee1 = request.POST.get('additional_fee1' + str(i), None)
        if additional_fee1 == '' : 
            additional_fee1 = None
        additional_fee2 = request.POST.get('additional_fee2' + str(i), None)
        if additional_fee2 == '' : 
            additional_fee2 = None
        additional_fee3 = request.POST.get('additional_fee3' + str(i), None)
        if additional_fee3 == '' : 
            additional_fee3 = None

        if productno != '':
            product_data = Product(productno, main_category, middle_category, sub_category, professionalism_fee_rate, 
                            potential_fee_rate, additional_fee1, additional_fee2, additional_fee3 ) 
            product_data.save()
            isSuccess = "성공"
        else :
            isSuccess = "실패"
    return isSuccess

def upload(request):
    isSuccess = "실패"
    if request.method == 'POST':
        if 'file' in request.FILES:
            wb = load_workbook(filename=request.FILES['file'].file)
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            print(worksheet)
            
            for row in worksheet.iter_rows(min_row=2): # Offset for header
                product = Product()
                product.productno = row[0].value
                product.main_category = row[1].value
                product.middle_category = row[2].value
                product.sub_category = row[3].value
                product.professionalism_fee_rate = row[4].value
                product.potential_fee_rate = row[5].value
                product.additional_fee1 = row[6].value
                product.additional_fee2 = row[7].value
                product.additional_fee3 = row[8].value

                if product.productno != '' and product.productno != None:
                    product.save()
                    isSuccess = "성공"
    return isSuccess

def download(request):
    isSuccess = "실패"
    # 워크북(엑셀파일)을 새로 만듭니다.
    wb = openpyxl.Workbook()

    # 현재 활성화된 시트를 선택합니다.
    sheet = wb.active
    # A1셀에 hello world!를 입력합니다.
    sheet['A1'] = '제품'

    # 워크북(엑셀파일)을 원하는 이름으로 저장합니다.

    i = 1
    sheet.cell(row=i, column=1).value = "Product NO"
    sheet.cell(row=i, column=2).value = "대분류"
    sheet.cell(row=i, column=3).value = "중분류"
    sheet.cell(row=i, column=4).value = "소분류"
    sheet.cell(row=i, column=5).value = "전문성 수수료율"
    sheet.cell(row=i, column=6).value = "잠재력 수수료율"
    sheet.cell(row=i, column=7).value = "추가 수수료1"
    sheet.cell(row=i, column=8).value = "추가 수수료2"
    sheet.cell(row=i, column=9).value = "추가 수수료3"
    i = i+1

    for data in Product.objects.all():
        sheet.cell(row=i, column=1).value = data.productno
        sheet.cell(row=i, column=2).value = data.main_category
        sheet.cell(row=i, column=3).value = data.middle_category
        sheet.cell(row=i, column=4).value = data.sub_category
        sheet.cell(row=i, column=5).value = data.professionalism_fee_rate
        sheet.cell(row=i, column=6).value = data.potential_fee_rate
        sheet.cell(row=i, column=7).value = data.additional_fee1
        sheet.cell(row=i, column=8).value = data.additional_fee2
        sheet.cell(row=i, column=9).value = data.additional_fee3
        i = i+1
    
    wb.save('제품.xlsx')

    return isSuccess
 