from .models import SamsungCode
from openpyxl import load_workbook
from openpyxl import Workbook

def save(request):

    print("samsung_code_save")

    for i in range(0,5):

        samsung_code = request.POST.get('samsung_code' + str(i), None)
        manager = request.POST.get('manager' + str(i), None)
        
        department = request.POST.get('department'+ str(i), None)
        if department == '' : 
            department = None
        phone_num = request.POST.get('phone_num'+ str(i), None)
        if phone_num == '' : 
            phone_num = None
        email = request.POST.get('email'+ str(i), None)
        if email == '' : 
            email = None

        if samsung_code != '' and manager != '' and samsung_code != None and manager != None:
            samsung_code_data = SamsungCode(samsung_code, manager, department, phone_num, email) 
            samsung_code_data.save()

    print("완료")

def upload(request):
    isSuccess = "실패"
    if request.method == 'POST':
        if 'file' in request.FILES:
            wb = load_workbook(filename=request.FILES['file'].file)
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            print(worksheet)
            
            for row in worksheet.iter_rows(min_row=2): # Offset for header
                samsung_code = SamsungCode()
                samsung_code.samsung_code = row[0].value
                samsung_code.manager = row[1].value
                samsung_code.department = row[2].value
                samsung_code.phone_num = row[3].value
                samsung_code.email = row[4].value

                if samsung_code.samsung_code != '' and samsung_code.manager != '' and samsung_code.samsung_code != None and samsung_code.manager != None:
                    samsung_code.save()
                    isSuccess = "성공"
    return isSuccess