from .models import Employee
from django.contrib.auth.models import User
from django.contrib import auth
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from django.contrib.auth import authenticate
def save(request):

    print("employee_save")

    for i in range(0,1):

        name = None; position = None; department = None;  phone_num = None;  office_num = None;  
        address = None; email = None; resident_registration_number = None; team = None; rate = None

        name = request.POST.get('name' + str(i), None)
        position = request.POST.get('position'+ str(i), None)
        if position == '' : 
            position = None
        department = request.POST.get('department'+ str(i), None)
        if department == '' : 
            department = None
        phone_num = request.POST.get('phone_num' + str(i), None)
        if phone_num == '' : 
            phone_num = None
        office_num = request.POST.get('office_num'+ str(i), None)
        if office_num == '' : 
            office_num = None
        address = request.POST.get('address'+ str(i), None)
        if address == '' : 
            address = None
        email = request.POST.get('email'+ str(i), None)
        if email == '' : 
            email = None
        resident_registration_number = request.POST.get('resident_registration_number'+ str(i), None)
        if resident_registration_number == '' : 
            resident_registration_number = None
        team = request.POST.get('team'+ str(i), None)
        if team == '' : 
            team = None
        rate = request.POST.get('rate'+ str(i), None)
        if rate == '' : 
            rate = None
        authority = request.POST.get('authority'+ str(i), None)
        if authority == '' : 
            authority = None
        id = request.POST.get('id'+ str(i), None)
        if id == '' : 
            id = None
        passwd = request.POST.get('passwd'+ str(i), None)
        if passwd == '' : 
            passwd = None
        charge = request.POST.get('charge'+ str(i), None)
        if charge == '' : 
            charge = None

        if name != '' and name != None:
            employee_data = Employee(name, team, position, department, resident_registration_number, rate, phone_num, office_num, address, email, charge, id, passwd, authority) 
            employee_data.save()
            print("성공1")
            isSuccess = "성공"
            if User.objects.all().filter(username=id).count() != 0:
                # 있으면
                user = User.objects.get(username=id)
                user.password = passwd
                user.email = authority
                user.save()
            else:
                # 없으면
                User.objects.create(username=id, password=passwd, first_name=id)
                user = User.objects.get(username=id)
                user.first_name=id
                user.email = authority
                user.set_password(passwd) # 비밀번호 변경 함수
                user.save()

            login_user = authenticate(username=id, password=passwd)
            print(login_user)
            if login_user is not None:
                print("등록 성공")
            else : 
                print("등록 실패")
        else:
            isSuccess = "실패"

    print("임직원 완료")
    return isSuccess

def upload(request):
    isSuccess = "실패"
    # try:
    if request.method == 'POST':
        if 'file' in request.FILES:
            wb = load_workbook(filename=request.FILES['file'].file)
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            print(worksheet)
            
            for row in worksheet.iter_rows(min_row=2): # Offset for header
                if(row[0].value == None):
                    break
                employee = Employee()
                employee.name = row[0].value
                employee.team = row[1].value
                employee.position = row[2].value
                employee.department = row[3].value
                employee.resident_registration_number = row[4].value
                employee.rate = row[5].value
                employee.phone_num = row[6].value
                employee.office_num = row[7].value
                employee.address = row[8].value
                employee.email = row[9].value
                employee.charge = row[10].value
                employee.id = row[11].value
                employee.passwd = row[12].value
                employee.authority = row[13].value
                
                if employee.name != '' and employee.name != None:
                    employee.save()
                    isSuccess = "성공"
                    if employee.id != '' and employee.id != None:
                        if User.objects.all().filter(username=employee.id).count() != 0:
                            # 있으면
                            user = User.objects.get(username=employee.id)
                            user.set_password(employee.passwd) # 비밀번호 변경 함수
                            user.email = employee.authority
                            user.save()
                        else:
                            # 없으면
                            User.objects.create(username=employee.id, password=employee.passwd, first_name=employee.id)
                            user = User.objects.get(username=employee.id)
                            user.first_name=employee.id
                            user.set_password(employee.passwd) # 비밀번호 변경 함수
                            user.email = employee.authority
                            user.save()

                            # login_user = authenticate(username=employee.id, password=employee.passwd)
                            # print(login_user)
                            # if login_user is not None:
                            #     print("등록 성공")
                            # else : 
                            #     print("등록 실패")
    # except:
    #     isSuccess = "실패"
    return isSuccess

def download(request):
    isSuccess = "성공"
    try:
        # 워크북(엑셀파일)을 새로 만듭니다.
        wb = openpyxl.Workbook()

        # 현재 활성화된 시트를 선택합니다.
        sheet = wb.active
        # A1셀에 hello world!를 입력합니다.
        sheet['A1'] = '임직원'

        # 워크북(엑셀파일)을 원하는 이름으로 저장합니다.

        i = 1
        sheet.cell(row=i, column=1).value = "이름"
        sheet.cell(row=i, column=2).value = "소속"
        sheet.cell(row=i, column=3).value = "직위"
        sheet.cell(row=i, column=4).value = "부서"
        sheet.cell(row=i, column=5).value = "주민등록번호"
        sheet.cell(row=i, column=6).value = "수수료율"
        sheet.cell(row=i, column=7).value = "휴대폰"
        sheet.cell(row=i, column=8).value = "내선번호"
        sheet.cell(row=i, column=9).value = "집주소"
        sheet.cell(row=i, column=10).value = "이메일"
        sheet.cell(row=i, column=11).value = "담당"
        sheet.cell(row=i, column=12).value = "id"
        sheet.cell(row=i, column=13).value = "passwd"
        sheet.cell(row=i, column=14).value = "권한"
        i = i+1

        for data in Employee.objects.all():
            sheet.cell(row=i, column=1).value = data.name
            sheet.cell(row=i, column=2).value = data.team
            sheet.cell(row=i, column=3).value = data.position
            sheet.cell(row=i, column=4).value = data.department
            sheet.cell(row=i, column=5).value = data.resident_registration_number
            sheet.cell(row=i, column=6).value = data.rate
            sheet.cell(row=i, column=7).value = data.phone_num
            sheet.cell(row=i, column=8).value = data.office_num
            sheet.cell(row=i, column=9).value = data.address
            sheet.cell(row=i, column=10).value = data.email
            sheet.cell(row=i, column=11).value = data.charge
            sheet.cell(row=i, column=12).value = data.id
            sheet.cell(row=i, column=13).value = data.passwd
            sheet.cell(row=i, column=14).value = data.authority
            i = i+1

        wb.save('임직원.xlsx')
    except:
        isSuccess = "실패"
    return isSuccess
 
