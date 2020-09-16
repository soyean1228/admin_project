from .models import OrderData
from .models import Approval
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl

def save(request):
    print("order_save")
    isSuccess = "저장에 실패했습니다"   

    # 저장 작업 전 데이터
    previous_data_origin = OrderData.objects.all()
    previous_data = []
    for i in previous_data_origin :
        previous_data.append(i)

    # 모든 데이터 삭제 
    queryset = OrderData.objects.all()
    queryset.delete() # 일괄 delete 요청 

    oppty_num = request.POST.get('oppty_num',0)
    quote_num = request.POST.get('select_quote_num',0)
    approval_data_length = request.POST.get('approval_data_length',0)

    # 저장 
    for i in range(0, int(approval_data_length) + 1):
        order_date = None;  order_num = None; assignment = None; 
        productno = None; buy_place = None; approval_quantity = None; order_quantity = None; 
        approval_unit = None;  approval_price = None; sales_unit = None; sales_price = None; 
        delivery_request_date = None; scheduled_delivery_date = None; recipient = None
        order_balance = None; 
        
        order_num = request.POST.get('order_num' + str(i),0)
        order_date = request.POST.get('order_date' + str(i),0)
        if order_date == '' : 
            order_date = None
        assignment = request.POST.get('assignment' + str(i),0)
        if assignment == '' : 
            assignment = None
        productno = request.POST.get('productno' + str(i),0)
        if productno == '' : 
            productno = None
        order_quantity = request.POST.get('order_quantity' + str(i),0)
        if order_quantity == '' : 
            order_quantity = None

        approval_price = request.POST.get('approval_price' + str(i),0)
        if approval_price == '' : 
            approval_price = None
        scheduled_delivery_date = request.POST.get('scheduled_delivery_date' + str(i),0)
        if scheduled_delivery_date == '' : 
            scheduled_delivery_date = None
        recipient = request.POST.get('recipient' + str(i),0)
        if recipient == '' : 
            recipient = None

        order_balance = order_quantity

        if order_num != '' :
            data_count = OrderData.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num, order_num=order_num).count()

            if data_count != 0 :
                print("동일한 데이터 존재")
                data = OrderData.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num, order_num=order_num)
                # approval의 approval_balance 구현
                approval_data = Approval.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num)
                approval_balance = approval_data.first().approval_balance + int(data.first().order_quantity)
                approval_data.update(approval_balance=approval_balance)
                data.delete()

            # approval의 approval_balance 구현
            approval_data = Approval.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num)
            approval_balance = approval_data.first().approval_balance - int(order_quantity)
            approval_data.update(approval_balance=approval_balance)

            OrderData.objects.create(quote_num=quote_num, oppty_num=oppty_num, productno=productno, recipient=recipient, order_num=order_num,order_date=order_date,assignment=assignment,
                                    order_quantity=order_quantity, scheduled_delivery_date=scheduled_delivery_date, order_balance=order_balance )

        isSuccess = "저장되었습니다"  

    # 저장 후 validation 체크 
    ## approval의 데이터에서 -1이 있으면 fail
    isFail = False 
    approval_data = Approval.objects.all()
    for i in approval_data :
        if int(i.approval_balance) < 0 :
            isSuccess = "주문수량은 승인수량보다 작아야 합니다"
            isFail = True

    #저장 실패하면 previous_data 복원
    if isFail == True :
        queryset = OrderData.objects.all()
        queryset.delete() # 일괄 delete 요청 

    for i in previous_data:
        OrderData.objects.create(quote_num=i.quote_num, oppty_num=i.oppty_num, productno=i.productno, recipient=i.recipient, order_num=i.order_num,order_date=i.order_date,assignment=i.assignment,
                                    order_quantity=i.order_quantity, scheduled_delivery_date=i.scheduled_delivery_date, order_balance=i.order_balance )
    # approval의 approval_balance 다시 계산
    approval_data = Approval.objects.all()
    for i in approval_data:
        order_data = OrderData.objects.filter(oppty_num=i.oppty_num, productno=i.productno, recipient=i.recipient, quote_num=i.quote_num)
        if order_data != None:
            order_quantity_sum = 0
            for j in order_data:
                order_quantity_sum = order_quantity_sum + j.order_quantity
            update_data = Approval.objects.filter(oppty_num=i.oppty_num,productno=i.productno,recipient=i.recipient, quote_num=i.quote_num)
            approval_balance = update_data.first().approval_quantity - order_quantity_sum
            update_data.update(approval_balance=approval_balance)

    return isSuccess
    
def upload(request):
    isSuccess = "실패"
    try:
        if request.method == 'POST':
            if 'file' in request.FILES:
                wb = load_workbook(filename=request.FILES['file'].file)
                first_sheet = wb.get_sheet_names()[0]
                worksheet = wb.get_sheet_by_name(first_sheet)
                print(worksheet)
                
                for row in worksheet.iter_rows(min_row=2): # Offset for header
                    if(row[0].value == None):
                        break
                    data = OrderData()
                    data.quote_num = row[2].value
                    if Approval.objects.filter(quote_num=row[2].value).count() != 0 :
                        approval = Approval.objects.filter(quote_num=row[2].value).first()
                    
                        data.oppty_num = approval.oppty_num
                        data.productno = approval.productno
                        data.recipient = approval.recipient

                        data.order_num = row[3].value
                        data.order_date = row[0].value
                        data.assignment = None
                        data.order_quantity = row[6].value
                        if row[15].value != None:
                            employee.scheduled_delivery_date = row[15].value
                        if row[16].value != None:
                            employee.scheduled_delivery_date = row[16].value

                        if data.order_num != '' and data.order_num != None:
                            data.save()
                            isSuccess = "성공" 
    except:
        isSuccess = "실패"
    return isSuccess

def modify(request):
    print("modify_order")
    modify_select_order_num = request.POST.get('modify_select_order_num','')
    modify_select_quote_num = request.POST.get('modify_select_quote_num','')
    modify_select_oppty_num = request.POST.get('modify_select_oppty_num','')
    modify_select_product_no = request.POST.get('modify_select_product_no','')
    modify_select_recipient = request.POST.get('modify_select_recipient','')
    order_data = OrderData.objects.raw('SELECT * FROM ( SELECT approval. *,  proposal.sales_unit, proposal.sales_price, proposal.delivery_address, proposal.recipient_phone1, proposal.recipient_phone2, proposal.buy_place, proposal.decision_quantity,  proposal.delivery_request_date  FROM  approval INNER JOIN proposal ON approval.oppty_num = proposal.oppty_num AND approval.productno = proposal.productno AND approval.recipient = proposal.recipient ) temp right JOIN order_data ON temp.productno=order_data.productno AND temp.oppty_num=order_data.oppty_num AND temp.recipient=order_data.recipient AND temp.quote_num=order_data.quote_num;')
    result_data = []
    for i in order_data :
        isInsert = 0
        if modify_select_order_num != '':
            if int(i.order_num) == int(modify_select_order_num):
                isInsert = 1
            else : 
                isInsert = 0 
                continue
        if modify_select_quote_num != '':
            if i.quote_num != None :
                if int(i.quote_num) == int(modify_select_quote_num):
                    isInsert = 1
                else : 
                    isInsert = 0 
                    continue
        if modify_select_oppty_num != '':
            if i.oppty_num == modify_select_oppty_num:
                isInsert = 1
            else : 
                isInsert = 0 
                continue
        if modify_select_product_no != '':
            if i.productno == modify_select_product_no:
                isInsert = 1
            else : 
                isInsert = 0 
                continue
        if modify_select_recipient != '':
            if i.recipient == modify_select_recipient:
                isInsert = 1
            else : 
                isInsert = 0 
                continue
        if isInsert == 1:
            result_data.append(i)
    return result_data

def modify_save(request):
    print("modify_save")

    # 수정 작업 전 데이터
    previous_data_origin = OrderData.objects.all()
    previous_data = []
    for i in previous_data_origin :
        previous_data.append(i)

    # 수정 전 해당 데이터 모두 삭제
    deleted_data = modify(request)
    order_data = OrderData.objects.raw('SELECT * FROM ( SELECT approval. *,  proposal.sales_unit, proposal.sales_price, proposal.delivery_address, proposal.recipient_phone1, proposal.recipient_phone2, proposal.buy_place, proposal.decision_quantity,  proposal.delivery_request_date  FROM  approval INNER JOIN proposal ON approval.oppty_num = proposal.oppty_num AND approval.productno = proposal.productno AND approval.recipient = proposal.recipient ) temp right JOIN order_data ON temp.productno=order_data.productno AND temp.oppty_num=order_data.oppty_num AND temp.recipient=order_data.recipient AND temp.quote_num=order_data.quote_num;')
    for i in order_data:
        for j in deleted_data:
            if i == j : 
                # approval의 approval_balance 구현
                approval_data = Approval.objects.filter(quote_num=i.quote_num, productno=i.productno, recipient=i.recipient, oppty_num=i.oppty_num)
                approval_balance_result = approval_data.first().approval_balance + int(i.order_quantity)
                approval_data.update(approval_balance=approval_balance_result)
                i.delete()

    isSuccess = "저장에 실패했습니다"  

    # 수정
    modify_select_data_length = request.POST.get('modify_select_data_length',0)
    for i in range(0, int(modify_select_data_length) + 1):
        order_date = None;  order_num = None; assignment = None; quote_num = None; 
        productno = None; buy_place = None; approval_quantity = None; order_quantity = None; 
        approval_unit = None;  approval_price = None; sales_unit = None; sales_price = None; 
        delivery_request_date = None; scheduled_delivery_date = None; recipient = None
        order_balance = None; oppty_num = None; 
        
        order_num = request.POST.get('modify_order_num' + str(i),'')
        quote_num = request.POST.get('modify_quote_num' + str(i),'')
        oppty_num = request.POST.get('modify_oppty_num' + str(i),'')
        productno = request.POST.get('modify_productno' + str(i),'')
        recipient = request.POST.get('modify_recipient' + str(i),'')

        order_date = request.POST.get('modify_order_date' + str(i),0)
        if order_date == '' : 
            order_date = None
        assignment = request.POST.get('modify_assignment' + str(i),0)
        if assignment == '' : 
            assignment = None
        order_quantity = request.POST.get('modify_order_quantity' + str(i),0)
        if order_quantity == '' : 
            order_quantity = None
        approval_quantity = request.POST.get('modify_approval_quantity' + str(i),0)
        if approval_quantity != '' :
            approval_quantity = None
        scheduled_delivery_date = request.POST.get('modify_scheduled_delivery_date' + str(i),0)
        if scheduled_delivery_date == '' : 
            scheduled_delivery_date = None

        order_balance = order_quantity

        if order_num != '' and quote_num != '':
            data_count = OrderData.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num, order_num=order_num).count()
            if data_count != 0 :
                print("삭제")
                data = OrderData.objects.get(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num, order_num=order_num)
                data.delete()
            else:
                # approval의 approval_balance 구현
                approval_data = Approval.objects.filter(quote_num=quote_num, productno=productno, recipient=recipient, oppty_num=oppty_num)
                approval_balance_result = approval_data.first().approval_balance - int(order_quantity)
                approval_data.update(approval_balance=approval_balance_result)

                OrderData.objects.create(quote_num=quote_num, oppty_num=oppty_num, productno=productno, recipient=recipient, order_num=order_num,order_date=order_date,assignment=assignment,
                                        order_quantity=order_quantity, scheduled_delivery_date=scheduled_delivery_date, order_balance=order_balance )
                
        isSuccess = "저장되었습니다"  
    
    # 수정 후 validation 체크 
    ## approval의 데이터에서 -1이 있으면 fail
    isFail = False 
    approval_data = Approval.objects.all()
    for i in approval_data :
        if int(i.approval_balance) < 0 :
            isSuccess = "주문수량은 승인수량보다 작아야 합니다"
            isFail = True

    #수정 실패하면 previous_data 복원
    if isFail == True :
        queryset = OrderData.objects.all()
        queryset.delete() # 일괄 delete 요청 
        print("복원")
        print(previous_data)
        for i in previous_data:
            OrderData.objects.create(quote_num=i.quote_num, oppty_num=i.oppty_num, productno=i.productno, recipient=i.recipient, order_num=i.order_num,order_date=i.order_date,assignment=i.assignment,
                                        order_quantity=i.order_quantity, scheduled_delivery_date=i.scheduled_delivery_date, order_balance=i.order_balance )
        # approval의 approval_balance 다시 계산
        approval_data = Approval.objects.all()
        for i in approval_data:
            order_data = OrderData.objects.filter(oppty_num=i.oppty_num, productno=i.productno, recipient=i.recipient, quote_num=i.quote_num)
            if order_data != None:
                order_quantity_sum = 0
                for j in order_data:
                    order_quantity_sum = order_quantity_sum + j.order_quantity
                update_data = Approval.objects.filter(oppty_num=i.oppty_num,productno=i.productno,recipient=i.recipient, quote_num=i.quote_num)
                approval_balance = update_data.first().approval_quantity - order_quantity_sum
                update_data.update(approval_balance=approval_balance)

    return isSuccess
