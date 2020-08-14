from .models import Approval
from .models import OrderData
from .models import Deposit
from .models import Delivery
from .models import Scm
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import datetime

def select(request):

    queryset = Scm.objects.all()
    queryset.delete() # 일괄 delete 요청    
    rawdata = Approval.objects.raw('SELECT tmp3.*, settlement.billing_date, settlement.billing_amount FROM ( SELECT tmp2.*, delivery.in_date, delivery.in_amount, delivery.delivery_date FROM ( SELECT tmp1.*,order_data.order_num, order_data.order_date, order_data.order_quantity, order_data.assignment, order_data.scheduled_delivery_date FROM ( SELECT proposal.*, approval.approval_quantity, approval.quote_num, approval.approval_unit, approval.approval_price FROM proposal LEFT JOIN approval ON proposal.oppty_num=approval.oppty_num AND proposal.productno=approval.productno AND proposal.recipient=approval.recipient) tmp1 LEFT JOIN order_data ON tmp1.oppty_num=order_data.oppty_num AND tmp1.productno=order_data.productno AND tmp1.recipient=order_data.recipient AND tmp1.quote_num=order_data.quote_num ) tmp2 LEFT JOIN delivery ON delivery.order_num=tmp2.order_num AND delivery.company_registration_number=tmp2.company_registration_number) tmp3 LEFT JOIN settlement ON settlement.order_num=tmp3.order_num AND settlement.quote_num=tmp3.quote_num AND settlement.oppty_num=tmp3.oppty_num AND settlement.productno=tmp3.productno AND settlement.recipient=tmp3.recipient;')
    
    for i in rawdata:
        scm = Scm()
        scm.scm_num = Scm.get_num(scm) + 1 
        print(scm.scm_num)
        scm.customer_name = i.customer_name
        scm.oppty_num = i.oppty_num
        scm.quote_num = i.quote_num
        scm.order_date = i.order_date
        scm.order_num = i.order_num
        scm.recipient = i.recipient
        scm.delivery_address = i.delivery_address
        scm.recipient_phone1 = i.recipient_phone1
        scm.recipient_phone2 = i.recipient_phone2
        scm.productno = i.productno
        scm.buy_place = i.buy_place
        scm.decision_quantity = i.decision_quantity
        scm.decision_unit = i.decision_unit
        scm.decision_price = i.decision_price
        scm.approval_unit = i.approval_unit
        scm.approval_price = i.approval_price
        scm.sales_unit = i.sales_unit
        scm.sales_price = i.sales_price
        scm.assignment = i.assignment
        scm.delivery_request_date = i.delivery_request_date
        scm.scheduled_delivery_date = i.scheduled_delivery_date
        scm.delivery_date = i.delivery_date
        scm.in_date = i.in_date
        scm.in_amount = i.in_amount
        scm.billing_place = i.billing_place
        scm.billing_date = i.billing_date
        scm.billing_amount = i.billing_amount
        scm.contact_conclusion_date = i.contact_conclusion_date
        scm.samsung_code = i.samsung_code
        scm.samsung_sales_manager = i.samsung_sales_manager
        scm.team = i.team
        scm.sales_manager = i.sales_manager
        scm.broker = i.broker
        scm.scm_manager = i.scm_manager
        scm.company_registration_number = i.company_registration_number
        scm.payment_method = i.payment_method
        scm.sales_type = i.sales_type
        scm.demand = i.demand
        scm.approval_quantity = i.approval_quantity
        scm.order_quantity = i.order_quantity
        scm.save()

    data = Scm.objects.all()
    print(data)
    for i in data:
        if i.order_num:
            i.decision_quantity = i.order_quantity
        elif i.quote_num:
            i.decision_quantity = i.approval_quantity
        i.save()
    
    return data

def find(request):

    data = Scm.objects.all()

    samsung_code = request.POST.get('samsung_code',0)
    if samsung_code != '':
        data = data.filter(samsung_code=samsung_code)
    samsung_sales_manager = request.POST.get('samsung_sales_manager',0)
    if samsung_sales_manager != '':
        data = data.filter(samsung_sales_manager=samsung_sales_manager)
    sales_manager = request.POST.get('sales_manager',0)
    if sales_manager != '':
        data = data.filter(sales_manager=sales_manager)
    scm_manager = request.POST.get('scm_manager',0)
    if scm_manager != '':
        data = data.filter(scm_manager=scm_manager)

    oppty_num = request.POST.get('oppty_num',0)
    if oppty_num != '':
        data = data.filter(oppty_num=oppty_num)
    quote_num = request.POST.get('quote_num',0)
    if quote_num != '':
        data = data.filter(quote_num=quote_num)
    order_num = request.POST.get('order_num',0)
    if order_num != '':
        data = data.filter(order_num=order_num)

    customer_name = request.POST.get('customer_name',0)
    if customer_name != '':
        data = data.filter(customer_name=customer_name)
    billing_place = request.POST.get('billing_place',0)
    if billing_place != '':
        data = data.filter(billing_place=billing_place)
    recipient = request.POST.get('recipient',0)
    if recipient != '':
        data = data.filter(recipient=recipient)
    delivery_address = request.POST.get('delivery_address',0)
    if delivery_address != '':
        data = data.filter(delivery_address=delivery_address)
    recipient_phone1 = request.POST.get('recipient_phone1',0)
    if recipient_phone1 != '':
        data = data.filter(recipient_phone1=recipient_phone1)

    order_date_start = request.POST.get('order_date_start','')
    order_date_end = request.POST.get('order_date_end','')

    delivery_request_date_start = request.POST.get('delivery_request_date_start',0)
    delivery_request_date_end = request.POST.get('delivery_request_date_end',0)

    scheduled_delivery_date_start = request.POST.get('scheduled_delivery_date_start' ,'')
    scheduled_delivery_date_end = request.POST.get('scheduled_delivery_date_end' ,'')

    delivery_date_start = request.POST.get('delivery_date_start',0)
    delivery_date_end = request.POST.get('delivery_date_end',0)
    
    result_data = []
    isInsert = 0
    for i in data :
        isInsert = 1
        if order_date_start != '' and order_date_end != '' :
            if i.order_date == None :
                isInsert = 0
            elif i.order_date.strftime('%Y-%m-%d') < order_date_start or i.order_date.strftime('%Y-%m-%d') > order_date_end: 
                isInsert = 0
        if delivery_request_date_start != '' and delivery_request_date_end != '':
            if i.delivery_request_date == None :
                isInsert = 0
            elif i.delivery_request_date.strftime('%Y-%m-%d') < delivery_request_date_start or i.delivery_request_date.strftime('%Y-%m-%d') > delivery_request_date_end: 
                isInsert = 0
        if scheduled_delivery_date_start != '' and scheduled_delivery_date_end != '':
            if i.scheduled_delivery_date == None :
                isInsert = 0
            elif i.scheduled_delivery_date.strftime('%Y-%m-%d') < scheduled_delivery_date_start or i.scheduled_delivery_date.strftime('%Y-%m-%d') > scheduled_delivery_date_end: 
                isInsert = 0
        if delivery_date_start != '' and delivery_date_end != '':
            if i.delivery_date == None :
                isInsert = 0
            elif i.delivery_date.strftime('%Y-%m-%d') < delivery_date_start or i.delivery_date.strftime('%Y-%m-%d') > delivery_date_end: 
                isInsert = 0
        if isInsert == 1:
            result_data.append(i)

    return result_data

def download(request):
    isSuccess = "엑셀 업로드에 성공했습니다"
    try:
        # 워크북(엑셀파일)을 새로 만듭니다.
        wb = openpyxl.Workbook()

        # 현재 활성화된 시트를 선택합니다.
        sheet = wb.active
        # A1셀에 hello world!를 입력합니다.
        sheet['A1'] = 'SCM 조회 결과'

        # 워크북(엑셀파일)을 원하는 이름으로 저장합니다.

        i = 1
        sheet.cell(row=i, column=1).value = "업체명"
        sheet.cell(row=i, column=2).value = "OPPTY번호"
        sheet.cell(row=i, column=3).value = "견적번호"
        sheet.cell(row=i, column=4).value = "주문일"
        sheet.cell(row=i, column=5).value = "주문번호"
        sheet.cell(row=i, column=6).value = "인수자"
        sheet.cell(row=i, column=7).value = "배송주소"
        sheet.cell(row=i, column=8).value = "인수자 전화번호1"
        sheet.cell(row=i, column=9).value = "Product No"
        sheet.cell(row=i, column=10).value = "매입처"
        sheet.cell(row=i, column=11).value = "수량"
        sheet.cell(row=i, column=12).value = "품의단가"
        sheet.cell(row=i, column=13).value = "품의액"
        sheet.cell(row=i, column=14).value = "승인단가"
        sheet.cell(row=i, column=15).value = "승인액"
        sheet.cell(row=i, column=16).value = "매출단가"
        sheet.cell(row=i, column=17).value = "매출액"
        sheet.cell(row=i, column=18).value = "할당"
        sheet.cell(row=i, column=19).value = "납품요청일"
        sheet.cell(row=i, column=20).value = "납품예정일"
        sheet.cell(row=i, column=21).value = "납품완료일"
        sheet.cell(row=i, column=22).value = "입금일"
        sheet.cell(row=i, column=23).value = "입금액"
        sheet.cell(row=i, column=24).value = "계산서발행처"
        sheet.cell(row=i, column=25).value = "계산서발행일"
        sheet.cell(row=i, column=26).value = "계산서발행금액"

        i = i+1

        for data in Scm.objects.all():
            sheet.cell(row=i, column=1).value = data.customer_name
            sheet.cell(row=i, column=2).value = data.oppty_num
            sheet.cell(row=i, column=3).value = data.quote_num
            if data.order_date != None:
                sheet.cell(row=i, column=4).value = data.order_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=4).value = None
            sheet.cell(row=i, column=5).value = data.order_num
            sheet.cell(row=i, column=6).value = data.recipient
            sheet.cell(row=i, column=7).value = data.delivery_address
            sheet.cell(row=i, column=8).value = data.recipient_phone1
            sheet.cell(row=i, column=9).value = data.productno
            sheet.cell(row=i, column=10).value = data.buy_place
            sheet.cell(row=i, column=11).value = data.decision_quantity
            sheet.cell(row=i, column=12).value = data.decision_unit
            sheet.cell(row=i, column=13).value = data.decision_price
            sheet.cell(row=i, column=14).value = data.approval_unit
            sheet.cell(row=i, column=15).value = data.approval_price
            sheet.cell(row=i, column=16).value = data.sales_unit
            sheet.cell(row=i, column=17).value = data.sales_price
            sheet.cell(row=i, column=18).value = data.assignment

            if data.delivery_request_date != None:
                sheet.cell(row=i, column=19).value = data.delivery_request_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=19).value = None

            if data.scheduled_delivery_date != None:
                sheet.cell(row=i, column=20).value = data.scheduled_delivery_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=20).value = None

            if data.delivery_date != None:
                sheet.cell(row=i, column=21).value = data.delivery_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=21).value = None

            if data.in_date != None:
                sheet.cell(row=i, column=22).value = data.in_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=22).value = None

            sheet.cell(row=i, column=23).value = data.in_amount
            sheet.cell(row=i, column=24).value = data.billing_place

            if data.billing_date != None:
                sheet.cell(row=i, column=25).value = data.billing_date.strftime('%Y-%m-%d')
            else :
                sheet.cell(row=i, column=25).value = None
            sheet.cell(row=i, column=26).value = data.billing_amount

            i = i+1

        wb.save('SCM 조회.xlsx')
    except:
        isSuccess = "엑셀 업로드에 실패했습니다"
    return isSuccess

 