from .models import Customer
from .models import Employee
from openpyxl import load_workbook
from openpyxl import Workbook

def save(request):
    print("customer_save")

    sales_charge_data = Employee.objects.filter(charge='INNO 영업담당')
    scm_charge_data = Employee.objects.filter(charge='INNO SCM담당')

    for i in range(0,1):
        customer_name = None; company_registration_number = None; representative = None; 
        address = None;  opening_date = None;  field1 = None; 
        field2 = None;  field3 = None;  sales_manager = None; 
        scm_manager = None;  rate = None;  

        company_registration_number = request.POST.get('company_registration_number', None)
        customer_name = request.POST.get('customer_name', None)
        if customer_name == '' : 
            customer_name = None
        representative = request.POST.get('representative', None)
        if representative == '' : 
            representative = None
        address = request.POST.get('address', None)
        if address == '' : 
            address = None
        opening_date = request.POST.get('opening_date' , None)
        if opening_date == '' : 
            opening_date = None
        field1 = request.POST.get('field1', None)
        if field1 == '' : 
            field1 = None
        field2 = request.POST.get('field2', None)
        if field2 == '' : 
            field2 = None
        field3 = request.POST.get('field3', None)
        if field3 == '' : 
            field3 = None
        purchasing_manager = request.POST.get('purchasing_manager', None)
        if purchasing_manager == '' : 
            purchasing_manager = None
        purchasing_contact_number1 = request.POST.get('purchasing_contact_number1' , None)
        if purchasing_contact_number1 == '' : 
            purchasing_contact_number1 = None
        purchasing_contact_number2 = request.POST.get('purchasing_contact_number2', None)
        if purchasing_contact_number2 == '' : 
            purchasing_contact_number2 = None
        purchasing_email = request.POST.get('purchasing_email' , None)
        if purchasing_email == '' : 
            purchasing_email = None
        settlement_manager = request.POST.get('settlement_manager' , None)
        if settlement_manager == '' : 
            settlement_manager = None
        settlement_contact_number = request.POST.get('settlement_contact_number' , None)
        if settlement_contact_number == '' : 
            settlement_contact_number = None
        settlement_email = request.POST.get('settlement_email' , None)
        if settlement_email == '' : 
            settlement_email = None
        sales_manager = request.POST.get('sales_manager', None)
        
        count_sales_charge_data = sales_charge_data.values('name').filter(name=sales_manager).count()
        print(count_sales_charge_data)
        if count_sales_charge_data == 0:
            return "영업 담당 오류"

        if sales_manager == '' : 
            sales_manager = None
        scm_manager = request.POST.get('scm_manager' , None)

        count_scm_charge_data = scm_charge_data.values('name').filter(name=scm_manager).count()
        print(count_scm_charge_data)
        if count_scm_charge_data == 0:
            return "SCM 담당 오류"

        if scm_manager == '' : 
            scm_manager = None
        rate = request.POST.get('rate' , None)
        if rate == '' : 
            rate = None

        if company_registration_number != '' :
            customer_data = Customer(customer_name, company_registration_number, representative, address, opening_date,
            field1, field2, field3, purchasing_manager, purchasing_contact_number1, purchasing_contact_number2, purchasing_email,
            settlement_manager, settlement_contact_number, settlement_email,
            sales_manager,  scm_manager, rate) 
            customer_data.save()
            isSuccess = "성공"
        else :
            isSuccess = "실패"
    return isSuccess


def upload(request):
    isSuccess = "실패"
    try:
        if request.method == 'POST':
            if 'file' in request.FILES:
                sales_charge_data = Employee.objects.filter(charge='INNO 영업담당')
                scm_charge_data = Employee.objects.filter(charge='INNO SCM담당')
                wb = load_workbook(filename=request.FILES['file'].file)
                first_sheet = wb.get_sheet_names()[0]
                worksheet = wb.get_sheet_by_name(first_sheet)
                print(worksheet)
                    
                for row in worksheet.iter_rows(min_row=2): # Offset for header
                    data = Customer()
                    data.customer_name = row[0].value
                    data.company_registration_number = row[1].value
                    data.representative = row[2].value
                    data.address = row[3].value
                    data.opening_date = row[4].value
                    data.field1 = row[5].value
                    data.field2 = row[6].value
                    data.field3 = row[7].value
                    data.purchasing_manager = row[8].value
                    data.purchasing_contact_number1 = row[9].value
                    data.purchasing_contact_number2 = row[10].value
                    data.purchasing_email = row[11].value
                    data.settlement_manager = row[12].value
                    data.settlement_contact_number = row[13].value
                    data.settlement_email = row[14].value
                    data.sales_manager = row[15].value
                    count_sales_charge_data = sales_charge_data.values('name').filter(name=data.sales_manager).count()
                    # print(count_sales_charge_data)
                    if count_sales_charge_data == 0:
                        return "영업 담당 오류"

                    data.scm_manager = row[16].value
                    count_scm_charge_data = scm_charge_data.values('name').filter(name=data.scm_manager).count()
                    # print(count_scm_charge_data)
                    if count_scm_charge_data == 0:
                        return "SCM 담당 오류"

                    data.rate = row[16].value
                    
                    if data.company_registration_number != '' and data.company_registration_number != None:
                        data.save()
                        isSuccess = "성공"
    except:
        isSuccess = "실패"
    return isSuccess

def download(request):
    isSuccess = "성공"
    try:
        # 워크북(엑셀파일)을 새로 만듭니다.
        wb = openpyxl.Workbook()

        # 현재 활성화된 시트를 선택합니다.
        sheet = wb.active
        # A1셀에 hello world!를 입력합니다.
        sheet['A1'] = '업체'

        # 워크북(엑셀파일)을 원하는 이름으로 저장합니다.

        i = 1
        sheet.cell(row=i, column=1).value = "업체명"
        sheet.cell(row=i, column=2).value = "사업자등록번호"
        sheet.cell(row=i, column=3).value = "대표자"
        sheet.cell(row=i, column=4).value = "주소"
        sheet.cell(row=i, column=5).value = "개업연월일"
        sheet.cell(row=i, column=6).value = "업종1"
        sheet.cell(row=i, column=7).value = "업종2"
        sheet.cell(row=i, column=8).value = "업종3"
        sheet.cell(row=i, column=9).value = "구매담당자"
        sheet.cell(row=i, column=10).value = "연락처1"
        sheet.cell(row=i, column=11).value = "연락처2"
        sheet.cell(row=i, column=12).value = "E-Mail"
        sheet.cell(row=i, column=13).value = "정산담당자"
        sheet.cell(row=i, column=14).value = "연락처"
        sheet.cell(row=i, column=15).value = "E-Mail"
        sheet.cell(row=i, column=16).value = "INNO 영업 담당"
        sheet.cell(row=i, column=17).value = "INNO SCM 담당"
        sheet.cell(row=i, column=18).value = "고객관리등급"
        i = i+1

        for data in Employee.objects.all():
            sheet.cell(row=i, column=1).value = data.customer_name
            sheet.cell(row=i, column=2).value = data.company_registration_number
            sheet.cell(row=i, column=3).value = data.representative
            sheet.cell(row=i, column=4).value = data.address
            sheet.cell(row=i, column=5).value = data.opening_date
            sheet.cell(row=i, column=6).value = data.field1
            sheet.cell(row=i, column=7).value = data.field2
            sheet.cell(row=i, column=8).value = data.field3
            sheet.cell(row=i, column=9).value = data.purchasing_manager
            sheet.cell(row=i, column=10).value = data.purchasing_contact_number1
            sheet.cell(row=i, column=11).value = data.purchasing_contact_number2
            sheet.cell(row=i, column=12).value = data.purchasing_email
            sheet.cell(row=i, column=13).value = data.settlement_manager
            sheet.cell(row=i, column=14).value = data.settlement_contact_number
            sheet.cell(row=i, column=15).value = data.settlement_email
            sheet.cell(row=i, column=16).value = data.sales_manager
            count_sales_charge_data = sales_charge_data.values('name').filter(name=data.sales_manager).count()
                    # print(count_sales_charge_data)
                    if count_sales_charge_data == 0:
                        return "영업 담당 오류"
            sheet.cell(row=i, column=17).value = data.scm_manager
            count_scm_charge_data = scm_charge_data.values('name').filter(name=data.scm_manager).count()
                    # print(count_scm_charge_data)
                    if count_scm_charge_data == 0:
                        return "SCM 담당 오류"
            sheet.cell(row=i, column=18).value = data.rate
            i = i+1
            
        wb.save('업체.xlsx')
    except:
        isSuccess = "실패"
    return isSuccess
