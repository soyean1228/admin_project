from .models import Broker
from .models import Employee
from openpyxl import load_workbook
from openpyxl import Workbook

def save(request):
    print("broker_save")

    for i in range(0,1):
        name = None; resident_registration_number = None; addresss = None; 
        contact_number = None; fee = None; team = None; manager = None

        name = request.POST.get('name' + str(i), None)
        resident_registration_number = request.POST.get('resident_registration_number' + str(i), None)
        if resident_registration_number == '' : 
            resident_registration_number = None
        addresss = request.POST.get('addresss' + str(i), None)
        if addresss == '' : 
            addresss = None
        contact_number = request.POST.get('contact_number' + str(i), None)
        if contact_number == '' : 
            contact_number = None
        fee = request.POST.get('fee' + str(i), None)
        if fee == '' : 
            fee = None
        team = request.POST.get('team' + str(i), None)
        if team == '' : 
            team = None
        manager = request.POST.get('manager', None)
        if manager == '' : 
            manager = None
        print(manager)
        count_manager = Employee.objects.all().values('name').filter(name=manager).count()
        print(count_manager)
        if count_manager == 0:
            return "담당자 오류"

        if name != '':
            broker_data = Broker(name, resident_registration_number, addresss, contact_number, fee, team, manager)
            broker_data.save()
            isSuccess = "성공"
        else :
            isSuccess = "실패"
    return isSuccess

def upload(request):
    isSuccess = "실패"
    if request.method == 'POST':
        if 'file' in request.FILES:
            wb = load_workbook(filename=request.FILES['file'].file)
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            print(worksheet)
            
            for row in worksheet.iter_rows(min_row=2): # Offset for header
                broker = Broker()
                broker.team = row[0].value
                broker.manager = row[1].value
                count_manager = Employee.objects.all().values('name').filter(name=broker.manager).count()
                print(count_manager)
                if count_manager == 0:
                    return "담당자 오류"
                broker.name = row[2].value
                broker.resident_registration_number = row[3].value
                broker.addresss = row[4].value
                broker.contact_number = row[5].value
                broker.fee = row[6].value

                if broker.name != '' and broker.name != None:
                    broker.save()
                    isSuccess = "성공"
    return isSuccess

def download(request):
    isSuccess = "성공"
    try:
        # 워크북(엑셀파일)을 새로 만듭니다.
        wb = openpyxl.Workbook()

        # 현재 활성화된 시트를 선택합니다.
        sheet = wb.active
        # A1셀에 hello world!를 입력합니다.
        sheet['A1'] = '알선사'

        # 워크북(엑셀파일)을 원하는 이름으로 저장합니다.

        i = 1
        sheet.cell(row=i, column=1).value = "성명(회사명)"
        sheet.cell(row=i, column=2).value = "주민(사업)번호"
        sheet.cell(row=i, column=3).value = "주소"
        sheet.cell(row=i, column=4).value = "연락처"
        sheet.cell(row=i, column=5).value = "소속"
        sheet.cell(row=i, column=6).value = "담당자"
        i = i+1

        for data in Employee.objects.all():
            sheet.cell(row=i, column=1).value = data.team
            sheet.cell(row=i, column=2).value = data.manager
            count_manager = Employee.objects.all().values('name').filter(name=broker.manager).count()
                print(count_manager)
                if count_manager == 0:
                    return "담당자 오류"
            sheet.cell(row=i, column=3).value = data.name
            sheet.cell(row=i, column=4).value = data.resident_registration_number
            sheet.cell(row=i, column=5).value = data.contact_number
            sheet.cell(row=i, column=6).value = data.fee
            i = i+1

        wb.save('알선사.xlsx')
    except:
        isSuccess = "성공"
    return isSuccess
